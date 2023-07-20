#-*- coding:utf-8 -*-
import codecs
import csv
import re
import os
import sys

#from natto import MeCab
import MeCab
import openpyxl

from typing import Tuple,List
from pathlib import Path
import numpy as np
import spacy
from sumy.summarizers.lex_rank import LexRankSummarizer

import pykakasi

nlp = spacy.load('ja_ginza')
kks = pykakasi.kakasi()

# 頻度辞書
nega_hindo_dic = {}
poji_hindo_dic = {}

__location__ = os.path.realpath(
    os.path.join(os.getcwd(), os.path.dirname(__file__)))

sample_filepath = "out.xlsx"
write_filepath = "kekka_ginza.xlsx"

# 1文の解析
# return
#
def sentense_ana(data,timestamp):

    docs = nlp(data)
    points = 0 # 文章全体の評価
    slot_dic = getSlotDic() # 回胴特有文言
    negaposi_dic = getNegaPosiDic() # 評価データの読み込み(さっき作った関数を呼び出している。)
    category_dic = getCategoryDic()
    hitei_dic = gethiteiDic()
    idiom_dic = getidiomDic() # 慣用句、連文
    #sentenses = re.split("[。！!♪♫★☆>?？（）w]", doc)  # 一文ごとに分ける
    category = 0
    category_type = 0

#    print(str(sentenses)) # 抽出ターゲット文言
    try:
        #######################################################################################################
        print("--------ネガポジ判定---------")
        for sentense in docs.sents: # 文の数だけ繰り返す
#            result_all = nm.parse(sentense) # 形態素解析して品詞分解をしている。
#            result_words = result_all.split("\n")  # 単語ごとに分ける
#            print(str(result_words)) #抽出ターゲット文言の形態素解析結果

            # 文言辞書
            token_dic=[]
            for token in sentense:
                token_dic.append(token)

            print(token_dic)
            for token in sentense:
                print(token.i, token.orth_, token.lemma_, token.pos_, token.dep_, token.tag_, token.head.i)
                negaposi = 0
                nega_flg = 0 #ネガポジワードだけど係り受け次第でネガポジが決まるワード用フラグ
                try:
                    if(token.i == "EOS"):
                        print("■■EOS■■")
                        break

                    ########################
                    # カテゴリ判定　※先決め
                    if(category == 0):
                        if token.orth_ in category_dic:
                            category = int(category_dic[token.orth_])  # その文のネガポジ
                            category_type = token.orth_

                    ########################
                    # 一般言語のネガポジ判定
                    #  係り受けと合わせて行う。
                    if token.lemma_ in negaposi_dic:
                        negaposi = int(negaposi_dic[token.lemma_])
                        if(negaposi != 0):  # その文のネガポジ
                            print("一般■" ,token.orth_, negaposi) # ネガポジ判定結果

                            #######
                            # ネガポジワードの係り受け処理
                            #  係り受けは動詞、形容詞に対して行う。
                            meisi_flg = 0
                            negaposi_2 = 0

                            ###############
                            # 特殊
                            #   係り受け次第でのネガワード検知用
                            #   例：もう、～ない
                            if(negaposi == 10):
                                negaposi = 0
                                nega_flg = 1
                                print("以降の文言でnegaposi判定")
                            else:
                                print("negaposiあり")

                            for moji in token_dic:

                                #########################
                                #
                                # １．親トークンの助動詞が係り受け
                                # ２．助動詞が親トークンの係り受け
                                #
                                if(token.i == moji.i):
                                    # 自分は対象外
                                    continue
                                # １．親トークンの助動詞が係り受け
                                if(token.head.i == moji.head.i):
                                    print("■■■対象文字■■■："+moji.lemma_)
                                    print("①親トークン　→　係り受け")

                                    #####
                                    # ひらがなに変換
                                    hiragana = kks.convert(moji.lemma_)

                                    print(hiragana[0]['hira'])
                                    # 自分含めて検索　必ず１つは見つかる
                                    if(hiragana[0]['hira'] in hitei_dic) :
                                        print("子トークンに否定を発見")
                                        if(negaposi != 0):
                                            if(negaposi_2 != 0):
                                                negaposi_2 = negaposi_2 * -1
                                            else:
                                                negaposi_2 = negaposi * -1
                                        else:
                                            negaposi_2 += -1

                                    elif(moji.pos_ == "NOUN" or moji.pos_ == "ADJ"or moji.pos_ == "AUX"):# or moji.pos_ == "VERB"):
                                        meisi_flg = 1
                                        print("名詞、助動詞、動詞発見")

                                        #########################
                                        # 親トークンが一致するものの中に否定があるかチェック
                                        for moji_2 in token_dic:
                                            if(moji.i == moji_2.i):
                                                # 自分は対象外
                                                continue
                                            if(moji.i == moji_2.head.i):
                                                #####
                                                # ひらがなに変換
                                                hiragana = kks.convert(moji_2.lemma_)

                                                print(hiragana[0]['hira'])
                                                # 自分含めて検索　必ず１つは見つかる
                                                if(hiragana[0]['hira'] in hitei_dic) :
                                                    print("子トークンに否定を発見")
                                                    if(negaposi != 0):
                                                        if(negaposi_2 != 0):
                                                            negaposi_2 = negaposi_2 * -1
                                                        else:
                                                            negaposi_2 = negaposi * -1
                                                    else:
                                                        negaposi_2 += -1
                                                print("negaposiワード："+str(negaposi_2))

                                    print("negaposiワード："+str(negaposi_2))

                                # ２．助動詞が親トークンの係り受け
                                elif(moji.head.i == token.i):
                                    """
                                    どっちか・・・
                                    elif(moji.i == token.head.i):　
                                    elif(moji.head.i == token.i):　もともと
                                    もともとの場合　②を検知してない？
                                    """
                                    print("■■■対象文字■■■："+moji.lemma_)
                                    print("②親トークン　←　係り受け")
                                    #####
                                    # ひらがなに変換
                                    hiragana = kks.convert(moji.lemma_)

                                    print(hiragana[0]['hira'])
                                    # 自分含めて検索　必ず１つは見つかる
                                    if(hiragana[0]['hira'] in hitei_dic) :
                                        print("子トークンに否定を発見")
                                        if(negaposi != 0):
                                            if(negaposi_2 != 0):
                                                negaposi_2 = negaposi_2 * -1
                                            else:
                                                negaposi_2 = negaposi * -1
                                        else:
                                            negaposi_2 += -1
                                    print("negaposiワード："+str(negaposi_2))

                                ########################
                                #
                                # 親トークンのかかっている文言が名詞かをチェック
                                #
                                if(token.head.i == moji.i):
                                    print("■■■対象文字■■■："+moji.lemma_)
                                    print("③親トークンが名詞に係り受け")
                                    if(moji.pos_ == "NOUN" or moji.pos_ == "ADJ"or moji.pos_ == "AUX" ):
                                        meisi_flg = 1
                                        print("名詞、助動詞発見２")

                                        #########################
                                        # 親トークンが一致するものの中に否定があるかチェック
                                        for moji_2 in token_dic:
                                            if(moji.i == moji_2.i):
                                                # 自分は対象外
                                                continue
                                            if(moji.head.i == moji_2.head.i):
                                                print("親トークン一致")

                                                #####
                                                # ひらがなに変換
                                                hiragana = kks.convert(moji_2.lemma_)

                                                print(hiragana[0]['hira'])
                                                # 自分含めて検索　必ず１つは見つかる
                                                if(hiragana[0]['hira'] in hitei_dic) :
                                                    print("子トークンに否定を発見")
                                                    if(negaposi != 0):
                                                        if(negaposi_2 != 0):
                                                            negaposi_2 = negaposi_2 * -1
                                                        else:
                                                            negaposi_2 = negaposi * -1
                                                    else:
                                                        negaposi_2 += -1
                                                print("negaposiワード："+str(negaposi_2))

                            ########################
                            # ネガポジ反転
                            #   ネガ意見にネガ意見が重なった時ポジ意見に反転する
                            #   ポジ意見にポジ意見が重なった時ネガ意見に反転する
                            if(negaposi_2 != 0):
                                if(nega_flg == 1): #そのまま入れる
                                    negaposi = negaposi_2
                                else:
                                    negaposi = negaposi * negaposi_2

                    ######################
                    # 回胴特有文言のネガポジ判定
                    if token.orth_ in slot_dic:
                        negaposi = int(slot_dic[token.orth_])  # その文のネガポジ
                        print("回胴特有■",token.orth_,negaposi)

                    points += negaposi # 文章全体の評価に加算

                    print("ネガポジ途中結果"+str(points))

                    ##########
                    # ■■■■　カテゴリ分類
#                EOFエラーかけたい
#                except EOFError:
#                    break
                except Exception as e:
                    print('%r' % e, flush=True)
            ##########
            # ■■■■文節でしか判断できない特殊文言
            # 即やめ


            ###############################
            # 慣用句、反語

            print("ターゲット■",sentense)
            renbun = str(sentense)
            # 慣用句、連文の抽出
            #renbun = (str)sentense
            for idiom in idiom_dic :
                if(renbun.find(idiom) >= 0):          # 対象の連文検知 : インデックス、非検知:-1
                    points += int(idiom_dic[idiom])   # ネガポジ結果を評価に加算
                    print("連文HIT■",idiom)

            # 反語表現対応
            if(renbun.find("?") >= 0):          # 対象の疑問符
                print("疑問符のため対象外■")
                points = 0   # 対象外
                category_type = 0
            elif(renbun.find("？") >= 0):       # 対象の疑問符
                points = 0   # 対象外
                print("疑問符のため対象外■")
                category_type = 0


            #############################
            # 頻度出し
            if(points > 0): # ポジの時
                for token in sentense:
                    if(token.pos_ == "NOUN" ):
                        if token.lemma_ in poji_hindo_dic:
                            poji_hindo_dic[token.lemma_] = int(poji_hindo_dic[token.lemma_]) + 1
                        else :
                            poji_hindo_dic[token.lemma_] = int(1)
            elif(points < 0): # ネガの時
                for token in sentense:
                    if(token.pos_ == "NOUN" ):
                        if token.lemma_ in nega_hindo_dic:
                            nega_hindo_dic[token.lemma_] = int(nega_hindo_dic[token.lemma_]) + 1
                        else :
                            nega_hindo_dic[token.lemma_] = int(1)

            #print(hindo_dic)
        #######################################################################################################

        print("-----------------")
        print("■カテゴリ:",category_type,category)
        print("■ネガポジ総計:",points)



    except Exception as e:
        print('%r' % e, flush=True)
        print(data, flush=True)


    ###
    # ■■■■　1件のカテゴリ、文、ネガポジ結果を返す。

    sentenses = re.split("[',]", data)  # 一文ごとに分ける
#    print(category)
#    print(sentenses)
    return category,points,timestamp,data # 文章全体の値を返す。

def gethiteiDic():
    with codecs.open(os.path.join(__location__, "./dataset/hitei.csv"), 'r', 'utf-8-sig') as f_in:
        reader = csv.reader(f_in, delimiter=',', lineterminator='\n')
        hiteiDic = {}
        for i, x in enumerate(reader):
            y = x[0].split(" ")
            hiteiDic[y[1]] = y[0]
    return hiteiDic

def getNegaPosiDic():
    with codecs.open(os.path.join(__location__, "./dataset/yougen.csv"), 'r', 'utf-8') as f_in:
        reader = csv.reader(f_in, delimiter=',', lineterminator='\n')
        negaPosiDic = {}
        for i, x in enumerate(reader):
            y = x[0].split(" ")
            negaPosiDic[y[1]] = y[0]
    with codecs.open(os.path.join(__location__, "./dataset/noun.csv"), 'r', 'utf-8') as f_in:
        reader = csv.reader(f_in, delimiter=',', lineterminator='\n')
        for i, x in enumerate(reader):
            y = x[0].split(" ")
            negaPosiDic[y[1]] = y[0]
    return negaPosiDic


#回胴特有表現
def getSlotDic():
    with codecs.open(os.path.join(__location__, "./dataset/slot.csv"), 'r', 'utf-8') as f_in:
        reader = csv.reader(f_in, delimiter=',', lineterminator='\n')
        slot_dic = {}
        for i, x in enumerate(reader):
            y = x[0].split(" ")
            slot_dic[y[1]] = y[0]
#    print(slot_dic)
    return slot_dic

#慣用句、連文
def getidiomDic():
    with codecs.open(os.path.join(__location__, "./dataset/idiom.csv"), 'r', 'utf-8-sig') as f_in:
        reader = csv.reader(f_in, delimiter=',', lineterminator='\n')
        idiom_dic = {}
        for i, x in enumerate(reader):
            y = x[0].split(" ")
            idiom_dic[y[1]] = y[0]
#    print(idiom_dic)
    return idiom_dic

#回胴カテゴリ
def getCategoryDic():
    with codecs.open(os.path.join(__location__, "./dataset/categori.csv"), 'r', 'utf-8') as f_in:
        reader = csv.reader(f_in, delimiter=',', lineterminator='\n')
        category_dic = {}
        for i, x in enumerate(reader):
            y = x[0].split(" ")
            category_dic[y[1]] = y[0]
    return category_dic

if __name__=='__main__':
    data = ""
    read_data = []
    #ここでリストの文言を一括読み込みして結果をみてみる
    #ファイル名一覧を作成する
    wb_gross = openpyxl.load_workbook(sample_filepath,data_only=True)
    ws_gross = wb_gross[wb_gross.sheetnames[0]]

    ###
    # リスト表題




    # ファイルの１行データを収集
    for row in ws_gross.iter_rows(min_row = 1):
        text = re.sub('[ 　「」『』【】！!♪♫★☆（）w]','', str(row[3].value)) # 不要な文言を削除 2021/12/02 add ?,？は除外
        read_data.append(sentense_ana(text,row[2].value))

    wb_new = openpyxl.Workbook()
    ws_new = wb_new[wb_new.sheetnames[0]]

    ws_new.append(["カテゴリ","ネガポジ","日付","コメント","日付(シリアル)"])
    for data in read_data:
        ws_new.append(data)
    #        ws_new.append(value_list)
    #        print(value_list)
    #        ws_new.append(value_list)

    ######
    # 頻度データ出力
    ws_new_sheet2 = wb_new.create_sheet(title='ネガ頻度')

    for hindo in nega_hindo_dic.items():
        ws_new_sheet2.append(hindo)

    ws_new_sheet3 = wb_new.create_sheet(title='ポジ頻度')

    for hindo in poji_hindo_dic.items():
        ws_new_sheet3.append(hindo)

    wb_new.save(write_filepath)
    """
        t = input()
        if t == '':
            break
        data = data + str(t)
    print(nlp(data))
    """
